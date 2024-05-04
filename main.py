import html
import random
import os , subprocess
import re
import shutil
import sys
from threading import Thread
import time
import requests
from bs4 import BeautifulSoup  
from urllib.parse import urljoin
from rich.console import Console
from rich.prompt import Prompt
from rich.text import Text
from rich.table import Table
from rich.layout import Layout
from utils import text_style,table_style
from rich.live import Live
from rich import box
from rich.progress import Progress
import keyboard
from openpyxl import Workbook, load_workbook
from rich.prompt import Confirm
from itertools import zip_longest
from zipfile import BadZipFile 


# initilizng things 
actions = ["View Recent Links","Scrap New Link"]
console = Console()
layout = Layout(name="root")
_width, _height = shutil.get_terminal_size()
console.size = (_width-1, _height-5)
page_counter = 0

meta_instructions = Text(justify="left")
meta_instructions.append("\U0001F44D Press SpaceBar to select\n",style="green")
meta_instructions.append("\U0001F3F9 Press Up & Down keys to nagivate\n",style="cyan")
meta_instructions.append("\U0001F645 Press Q to go back\n",style="red")
meta_instructions.append("\U0001F645 Press CTRL+C to exit\n",style=text_style.info_style)
base_domain = "http://can't be exposed sorry.com"
target_url = f"{base_domain}/toc/10991476/2024/47/6"

# Set headers to mimic a browser request
headers = {
    "Sec-Ch-Ua":'"Google Chrome";v="123", "Not:A-Brand";v="8", "Chromium";v="123"',
    "Sec-Ch-Ua-Mobile":"?0",
    "Sec-Ch-Ua-Platform": '"Windows"',
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
        }



# ---------------------Scraper functions and their helpers --------------------
# this function scrap the previous page link and the articals links of the current page provided 
def scrape_site(url):
    print(url)
    try:
        global page_counter
        page_counter +=1
        # Make an HTTP GET request to the URL with headers
        response = requests.post(url, headers=headers)
        response.raise_for_status()  # Raise an exception for bad status codes

        # Parse the HTML content
        soup = BeautifulSoup(response.content, 'html.parser')

        # Find all anchor tags with class "issue-item__title" which are articals links
        anchor_tags = soup.find_all('a', class_='issue-item__title')
        # Extract href values from anchor tags and add base domain
        urls = [urljoin(base_domain, tag['href']) for tag in anchor_tags]
        # droping first because thats not reasearch artical but issue 
        urls.pop(0)

    except Exception as e:
        console.print(f"Error while fetching articals from page: {e}",style=text_style.error_style)
        urls = None
    try:
         # Finding one and only anchor tag with class content-navigation__btn--pre which is previous page
        previous_tag = urljoin(base_domain,soup.find("a",class_="content-navigation__btn--pre")["href"])
    except:
        previous_tag = None
    return previous_tag, urls


def save_soup_to_file(soup, filename):
    try:
        with open(filename, "w", encoding="utf-8") as file:
            file.write(soup.prettify())
        console.print(f"Soup data saved to {filename}")
    except Exception as e:
        console.print(f"Error saving soup data to file: {e}")
def decode_email(encoded_string):
     # Extract the encoded part after the #
    match = re.search(r"#([0-9a-fA-F]+)$", encoded_string)
    if match:
        encoded_email = match.group(1)
    else:
        print("Error: No encoded email found.")
        return "Not available"
    try:
        # Decode the encoded email address
        decoded_email = ""
        r = int(encoded_email[:2], 16)
        for i in range(2, len(encoded_email), 2):
            decoded_email += chr(int(encoded_email[i:i+2], 16) ^ r)

        # Attempt to decode UTF-8
        decoded_email = html.unescape(decoded_email)
        return decoded_email
    except Exception as e:
        console.print("Error decoding email:", e)
        return None
def _validateScrapedData(author_name, email, address):
    # adjusting their lenght to fit in table 
    if(len(author_name) > 20):
        author_name = author_name[:20]+".."
    if(len(address) > 20):
        address = address[:20] + ".."
    if(len(email) >20):
        email = email[:20] + ".."
    return {"author_name":author_name,"email":email,"address":address}
def _storeRecordInExcel(author_name, email, address):
    # Check if data.xlsx exists in the current directory
    if os.path.exists("data.xlsx"):
        try:
            # If the file exists, load it
            wb = load_workbook("data.xlsx")
            ws = wb.active
        except BadZipFile:
            raise FileNotFoundError("Error: The file 'data.xlsx' appears to be corrupted or not in the expected format, Please delete the file first.")
    else:
        # If the file doesn't exist, create a new workbook and worksheet
        wb = Workbook()
        ws = wb.active
        # Add column headers
        ws.append(["Author Name", "Email", "Address"])

    # Append data to the worksheet
    ws.append([author_name, email, address])

    # Save the workbook to data.xlsx
    wb.save("data.xlsx")

def _save_and_validate_authors(authors):
    valid_authors = []
    for author in authors:
        _storeRecordInExcel(author["author_name"],author["email"],author["address"])
        valid_authors.append(_validateScrapedData(author["author_name"],author["email"],author["address"]))
    return valid_authors

def scrape_author_info(url):
    try:
        # Make an HTTP GET request to the URL
        response = requests.post(url,headers=headers)
        response.raise_for_status()  # Raise an exception for bad status codes

        # Parse the HTML content
        soup = BeautifulSoup(response.content, 'html.parser')
        # Find author's information
        parent = soup.find("div",id="sb-1").find("div",class_="comma__list").find("div","accordion-tabbed")
        span_childs = parent.findAll("span",recursive=False)
        only_authors_spans = [span for span in span_childs if span.find("i",class_="icon-mail_outline")]
        authors = []
        for count,current_span in enumerate(only_authors_spans):
            author = {}
            try:
                name = current_span.find("i",class_="icon-mail_outline").parent.text
            except Exception:
                name = "Not-Available"

            try:
                email_href = current_span.find("a", title="Link to email address")["href"]
                email = decode_email(email_href)
            except Exception:
                email = "Not-Available"
            try:
                address_p = current_span.find('p', class_='author-type',string='Corresponding Author').findNextSiblings("p")[1:]
                address_text_list = [p.get_text().strip() for p in address_p if not p.findChild("b") and not p.findChild("a") and not p.text.startswith("Communicated by:")]
                addressString = ' | '.join(address_text_list)
            except Exception:
                try:
                    address_p = current_span.find("p",class_="author-name").findNextSiblings("p")
                    address_text_list = [p.get_text().strip() for p in address_p if not p.findChild("b") and not p.findChild("a") and not p.text.startswith("Communicated by:")]
                    addressString = ' | '.join(address_text_list)
                except Exception:
                    addressString = "Not-Available"
            
            # adding all attributes to the object 
            author["author_name"] = name
            author["email"] = email
            author["address"] = addressString
            authors.append(author)
        return _save_and_validate_authors(authors)
    except requests.exceptions.RequestException as e:
        console.print(f"Error while fetching author details from artical: {e}",style=text_style.error_style)
        return {'author_name': "not-available","email":"not-available", 'address': "not-available"}
# --------------------------utility functions ---------------------
def completeThread(given_thread:Thread,thread_function):
    console.clear()
    console.clear_live()
    # starting given thred 
    given_thread.start()
    # waiting for its completion 
    given_thread.join()
    # returning new thread because previous will be dead 
    return Thread(target=thread_function)

def _validateLink(url):
    regex = r"^(http|https)://[^\s/$.?#].[^\s]*$"
    if re.match(regex, url):
        if url.startswith("https://"):
            url = url.replace("https://", "http://", 1)
        return url
    else:
        return None

def _take_new_link():
    while True:
        link = Prompt.ask("Enter Link: ", default=target_url)
        if link:
            return link
        else:
            console.print("Invalid URL. Please enter a valid URL.",style=text_style.error_style)
def _check_recent_links():
    # Check if recent.txt exists in the current directory
    if os.path.exists("recent.txt"):
        # Open the file for reading
        with open("recent.txt", "r") as file:
            # Read all lines from the file
            lines = file.readlines()
            # Initialize an empty list to store links
            links = []
            # Iterate over each line
            for line in lines:
                # Strip whitespace characters from the beginning and end of the line
                line = line.strip()
                # Check if the line starts with "http://" or "https://" (assuming links start with these)
                if line.startswith("http://") or line.startswith("https://"):
                    # Append the link to the list
                    links.append(line)
            # Return the list of links
            return links
    else:
        console.print("No recent links found",style=text_style.info_style)
        return []
# this function is to display recent saved links 
def _display_list_menu(items_list:list,sub_layout:Layout,active_row = 0):
    # adding additional info at the front 
    sub_layout.split_column(Layout(name="table_menu"))
    def update_table(current_row):
        menu_table = Table(box=box.ASCII_DOUBLE_HEAD,title="Recent Links",title_style=text_style.success_style,show_lines=True,show_header=False,style=table_style.default,expand=True,row_styles=["cyan"])
        for index,item in enumerate(items_list):
            if index == current_row:
                menu_table.add_row(f"{index+1}. {item}",style=text_style.active_row_style)
            else:
                menu_table.add_row(f"{index+1}. {item}")
        return menu_table

    with Live(layout,refresh_per_second=8) as live:  # update 4 times a second to feel fluid
        while True:
            time.sleep(0.08)  # arbitrary delay of 80 milli seconds
            if keyboard.is_pressed("down") and active_row < len(items_list) - 1: 
                active_row += 1
            elif keyboard.is_pressed("up") and active_row > 0: 
                active_row -= 1
            elif keyboard.is_pressed("space"): 
                live.stop()
                # returning selected link 
                return items_list[active_row]
            elif keyboard.is_pressed("Q"): 
                live.stop()
                main()
            sub_layout["table_menu"].update(update_table(active_row))

def _show_authors_details(page_url,articals):
    head_text = Text(f"Current Page Url: {page_url}\n",justify="center")
    head_text.append(f"Page no. {page_counter}",style="magenta bold")
    head_text.stylize("bold magenta",len(page_url)+18,len(page_url)+27)
    head_text.stylize("bold yellow",len(page_url)+27,None)
    head_text.stylize("bold magenta",0,17)
    head_text.stylize("bold yellow",17,len(page_url)+18)
    body_loading_text = Text(f"Please wait while gathering page's meta data...",justify="center",style=text_style.success_style)
    auther_details_layout = Layout(name="auther_details_layout")
    auther_details_layout.split_column(Layout(head_text,name="auther_head",size=2),Layout(body_loading_text,name="auther_body"),Layout(Text("JUFFLER says, enjoy scrapping \U0001F600",style=text_style.success_style, justify="center"),name="auther_footer",size=2))
    # setting head 
    console.clear()
    data_table = Table(box=box.ROUNDED,title=f"Authors details",title_style=text_style.success_style,show_lines=True,style="magenta",expand=True,row_styles=["cyan"],header_style=text_style.highlight_style)
    data_table.add_column("Name",justify="left")
    data_table.add_column("Email",justify="left")
    data_table.add_column("Address",justify="left")
    
    # Create a progress instance
    progress = Progress()
    # Add a task to the progress instance
    task = progress.add_task("[green]Processing...", total=len(articals))
    
    # Showing auther details 
    with Live(auther_details_layout,refresh_per_second=4) as live:
        for key, url in enumerate(articals):
            
            time.sleep(0.09)  # arbitrary delay
            try:
                # resetting table when 10 entries are filled 
                if(int(data_table.row_count)%10 == 0):
                    data_table = Table(box=box.ROUNDED,title=f"Authers details",title_style=text_style.success_style,show_lines=True,style="magenta",expand=True,row_styles=["cyan"],header_style=text_style.highlight_style)
                    data_table.add_column("Name",justify="left")
                    data_table.add_column("Email",justify="left")
                    data_table.add_column("Address",justify="left")
                else:
                    pass
                author_info = scrape_author_info(url)
                different_color = random.choice(text_style.colors)
                if author_info:
                    for current_author in author_info:
                        # if artica has more then 1 author then show in difference color
                        if(len(author_info) > 1):
                            data_table.add_row(f"[{different_color}]{key+1}.[/] {current_author['author_name']}",current_author["email"],current_author["address"])
                        else:
                            data_table.add_row(f"[yellow]{key+1}.[/] {current_author['author_name']}",current_author["email"],current_author["address"])
                else:
                    console.print("Failed to retrieve author information.")
                # rendering new data 
                auther_details_layout["auther_body"].update(data_table)
                auther_details_layout["auther_footer"].update(progress.get_renderable())
            except Exception as e:
                console.print(f"Unable to scrap URL #{key+1}: {url}")
            # updating progress 
            progress.update(task, advance=1)

# ------------------pages --------------------
def renderRecentLinksPage():
    recent_links = _check_recent_links()
    if(len(recent_links)!= 0):
        selected_link = _display_list_menu(recent_links,layout["body"])
        # running renderScrapNewLinkPage function in new terminal 
        command = [sys.executable, "main.py","-l",selected_link]
        subprocess.Popen(command, creationflags=subprocess.CREATE_NEW_CONSOLE,startupinfo=subprocess.STARTUPINFO())
        # making new thread for the main window 
        new_recent_link_page_thread = Thread(target=renderRecentLinksPage)
        new_recent_link_page_thread.start()
        new_recent_link_page_thread.join()
    else:
        main(Text("Alert! : No recent link found",style=text_style.error_style))

# this will recursively call it self untils previus link is not available 
def scrapAndShowArticals(page_link):
    previous_tag, articals = scrape_site(page_link)
    if articals:
        _show_authors_details(page_link,articals)
    else:
        console.print(f"Articals not found in page url: [yellow bold]{page_link}[/]",style=text_style.error_style)
    if(previous_tag):
        scrapAndShowArticals(previous_tag)
    
def _addLinkToRecents(link):
    recent_links = _check_recent_links()
    if(link not in recent_links):
        with open("recent.txt", "a") as file:
            file.write(link + "\n")
            
def renderScrapNewLinkPage(link = None):
    console.clear()
    if not link:
        link = _take_new_link()
    link = _validateLink(link)
    _addLinkToRecents(link)
    if link:
        scrapAndShowArticals(link)
    else:
        main(Text("Alert! link is not valid",style=text_style.error_style))

# ---------------making pages threads -----------
scrap_new_link_page_thread = Thread(target=renderScrapNewLinkPage)

def main(alert:Text = ""):
    try:
        # resetting console and initilizing utilit vars
        console.clear()
        console.clear_live()
        # recent page thread 
        recent_link_page_thread = Thread(target=renderRecentLinksPage)

        # start making layouts 
        layout.split_column(Layout(name="head",size=5,),Layout(name="body",ratio=1),Layout(name="footer",size=2))
        # Setting static header and footer UI's 
        layout["head"].split_column(Layout(name="greeting"),Layout(name="instructions"))
        greeting_text = Text("Welcome to the Site-Scrapper \U0001F600\n",style=text_style.info_style,justify="center")
        # adding alert to the heading is availble 
        if isinstance(alert,Text):
            greeting_text.append(alert)
        else:
            pass
        layout["head"]["greeting"].update(greeting_text)
        layout["head"]["instructions"].update(meta_instructions)
        layout["footer"].update(Text("Copyright 2024-25 \u00A9 reserved by JUFFLER",style=text_style.info_style,justify="center"))

        # making main menu 
        def update_main_menu(active_option):
            main_menu_table = Table(box=box.ROUNDED,title="Available actions",title_style=text_style.success_style,show_lines=True,show_header=False,style=table_style.default,expand=True,row_styles=["cyan"])
            for index,action in enumerate(actions):
                if index == active_option:
                    main_menu_table.add_row(f"{index+1}. {action}",style=text_style.active_row_style)
                else:
                    main_menu_table.add_row(f"{index+1}. {action}")
            return main_menu_table
        # final rendering of UI's
        with Live(layout, refresh_per_second=4) as live:  # update 4 times a second to feel fluid
            active_option=0
            while True:
                time.sleep(0.09)  # arbitrary delay
                # measuring keystoks 
                if keyboard.is_pressed("down") and active_option < len(actions) - 1: 
                    active_option += 1
                elif keyboard.is_pressed("up") and active_option > 0: 
                    active_option -= 1
                elif keyboard.is_pressed("space"):
                    live.stop()
                    if(active_option == 0):
                        recent_link_page_thread = completeThread(recent_link_page_thread,renderRecentLinksPage)
                    elif(active_option == 1):
                        command = [sys.executable, "main.py","scrap-link"]
                        subprocess.Popen(command, creationflags=subprocess.CREATE_NEW_CONSOLE,startupinfo=subprocess.STARTUPINFO())
                        main()
                    else:
                        console.print("Option not Available",style=text_style.error_style)
                        main()
                layout["body"].update(update_main_menu(active_option))
    except Exception as e:
        # print(e)
        console.print("Some error occure. Please wait ", style=text_style.error_style)

    
def handleSysArgument(args):
    if len(args) < 1:
        return 
    else:
        if("scrap-link" in args):
            scrap_new_link_page_thread.start()
            scrap_new_link_page_thread.join()
            # Close the terminal window using taskkill
            confirmation_text = Text("Task Completed, Enter y or n to close",style=text_style.success_style,justify="center")
            if(Confirm.ask(confirmation_text)):
                subprocess.run(["taskkill", "/F", "/FI", "PID eq {}".format(os.getpid())])
        elif("-l" in args):
            link = args[args.index("-l")+1]
            renderScrapNewLinkPage(link)
    
if __name__ == "__main__":
    # handle system arguments if available 
    handleSysArgument(sys.argv[1:])
    main()